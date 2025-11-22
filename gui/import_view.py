"""
Vista del módulo de Importación CSV
Autor: Joaquin Armando Loaiza Cruz
Fecha: 2025-11-11
Descripción: Importación de datos desde archivos CSV
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import Optional, Dict, Any
import csv
import sys
import os
from datetime import datetime, time

# Importar openpyxl para manejo de Excel
try:
    import openpyxl
except ImportError:
    openpyxl = None

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.attendance_service import AttendanceService
from database.employee_service import EmployeeService


class ImportView:
    """Vista para importar datos desde CSV y Excel"""
    
    def __init__(self, parent_frame: tk.Frame, attendance_service: Optional[AttendanceService] = None, employee_service: Optional[EmployeeService] = None):
        """
        Inicializa la vista de importación.
        
        Args:
            parent_frame: Frame contenedor principal
            attendance_service: Servicio de asistencias para importar datos
            employee_service: Servicio de empleados para validaciones o importación
        """
        self.parent_frame = parent_frame
        self.attendance_service = attendance_service
        self.employee_service = employee_service
        
    def render(self):
        """Renderiza la vista completa de importación"""
        # Limpiar frame padre
        for widget in self.parent_frame.winfo_children():
            widget.destroy()
        
        # Título
        self._create_header()
        
        # Contenedor principal
        main_container = tk.Frame(self.parent_frame, bg='#f5f5f5')
        main_container.pack(fill='both', expand=True, padx=30, pady=(0, 30))
        
        # Información
        info_frame = tk.Frame(main_container, bg='#e3f2fd', relief='solid', borderwidth=1)
        info_frame.pack(fill='x', pady=(0, 20))
        
        tk.Label(
            info_frame,
            text="💡 Información",
            font=('Segoe UI', 11, 'bold'),
            bg='#e3f2fd',
            fg='#1976d2',
            padx=20,
            pady=10
        ).pack(anchor='w')
        
        tk.Label(
            info_frame,
            text="Esta funcionalidad permite importar datos masivos desde archivos Excel y CSV.\n"
                 "Asegúrese de que el archivo tenga el formato correcto antes de importar.",
            font=('Segoe UI', 9),
            bg='#e3f2fd',
            fg='#1565c0',
            justify='left',
            padx=20
        ).pack(anchor='w', pady=(0, 15))
        
        # Opciones de importación
        self._create_import_options(main_container)
    
    def _create_header(self):
        """Crea el encabezado"""
        title_frame = tk.Frame(self.parent_frame, bg='white', height=80)
        title_frame.pack(fill='x', padx=30, pady=(30, 20))
        title_frame.pack_propagate(False)
        
        title_content = tk.Frame(title_frame, bg='white')
        title_content.pack(side='left', expand=True)
        
        tk.Label(
            title_content,
            text="📁",
            font=('Segoe UI', 28),
            bg='white'
        ).pack(side='left', padx=(0, 15))
        
        tk.Label(
            title_content,
            text="Importación de Datos (Excel/CSV)",
            font=('Segoe UI', 20, 'bold'),
            bg='white',
            fg='#2c3e50'
        ).pack(side='left')
    
    def _create_import_options(self, parent):
        """Crea las opciones de importación"""
        # Grid de opciones
        grid_frame = tk.Frame(parent, bg='#f5f5f5')
        grid_frame.pack(fill='both', expand=True)
        
        options = [
            {
                'title': '📝 Importar Asistencias (Excel)',
                'description': 'Importa registros desde "Reporte_AsistenciaDetallado.xlsx"',
                'columns': 'Hoja: Resumen Detallado (Codigo, Fecha, Turno, Ingreso, Salida)',
                'example': 'Formato estándar del reporte de asistencia',
                'command': lambda: self._import_attendance_excel(),
                'btn_text': '📂 Seleccionar Excel (.xlsx)'
            },
            {
                'title': '👥 Importar Empleados (Excel)',
                'description': 'Importa un listado de empleados desde Excel',
                'columns': 'Código, Nombre, DNI, Puesto, Centro Coste, Subdivisión',
                'example': 'E00001, Juan Pérez, 12345678, Operario, CC001, SUB1',
                'command': lambda: self._import_employees_excel(),
                'btn_text': '📂 Seleccionar Excel (.xlsx)'
            }
        ]
        
        for idx, option in enumerate(options):
            row = idx // 2
            col = idx % 2
            
            self._create_import_card(
                grid_frame,
                option['title'],
                option['description'],
                option['columns'],
                option['example'],
                option['command'],
                option['btn_text'],
                row,
                col
            )
    
    def _create_import_card(self, parent, title, description, columns, example, command, btn_text, row, col):
        """Crea una tarjeta de opción de importación"""
        card = tk.Frame(parent, bg='white', relief='solid', borderwidth=1)
        card.grid(row=row, column=col, padx=10, pady=10, sticky='nsew')
        
        # Configurar grid
        parent.grid_rowconfigure(row, weight=1)
        parent.grid_columnconfigure(col, weight=1)
        
        # Contenido
        content = tk.Frame(card, bg='white', padx=25, pady=20)
        content.pack(fill='both', expand=True)
        
        # Botón (Empaquetado al final con side=bottom para asegurar visibilidad)
        tk.Button(
            content,
            text=btn_text,
            command=command,
            font=('Segoe UI', 9, 'bold'),
            bg='#2196f3',
            fg='white',
            activebackground='#1976d2',
            relief='flat',
            cursor='hand2',
            padx=20,
            pady=10
        ).pack(side='bottom', fill='x', pady=(15, 0))
        
        # Frame superior para el resto del contenido
        top_content = tk.Frame(content, bg='white')
        top_content.pack(side='top', fill='both', expand=True)
        
        # Título
        tk.Label(
            top_content,
            text=title,
            font=('Segoe UI', 13, 'bold'),
            bg='white',
            fg='#2c3e50'
        ).pack(anchor='w', pady=(0, 10))
        
        # Descripción
        tk.Label(
            top_content,
            text=description,
            font=('Segoe UI', 9),
            bg='white',
            fg='#546e7a',
            wraplength=350,
            justify='left'
        ).pack(anchor='w', pady=(0, 15))
        
        # Formato esperado
        format_frame = tk.Frame(top_content, bg='#f5f5f5', relief='solid', borderwidth=1)
        format_frame.pack(fill='x')
        
        tk.Label(
            format_frame,
            text="Columnas:",
            font=('Segoe UI', 8, 'bold'),
            bg='#f5f5f5',
            fg='#37474f',
            anchor='w',
            padx=10,
            pady=5
        ).pack(fill='x')
        
        tk.Label(
            format_frame,
            text=columns,
            font=('Consolas', 8),
            bg='#f5f5f5',
            fg='#2c3e50',
            anchor='w',
            padx=10,
            pady=5
        ).pack(fill='x')
        
        tk.Label(
            format_frame,
            text="Ejemplo:",
            font=('Segoe UI', 8, 'bold'),
            bg='#f5f5f5',
            fg='#37474f',
            anchor='w',
            padx=10,
            pady=5
        ).pack(fill='x')
        
        tk.Label(
            format_frame,
            text=example,
            font=('Consolas', 8),
            bg='#f5f5f5',
            fg='#1976d2',
            anchor='w',
            padx=10
        ).pack(fill='x', pady=(0, 5))
    
    def _import_employees_excel(self):
        """Importa empleados desde Excel"""
        if not self.employee_service:
            messagebox.showerror("Error", "No hay conexión a la base de datos")
            return

        if openpyxl is None:
            messagebox.showerror("Error", "La librería 'openpyxl' no está instalada.")
            return

        filename = filedialog.askopenfilename(
            title="Seleccionar archivo Excel de empleados",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.load_workbook(filename, data_only=True)
            sheet = wb.active # Usar la hoja activa por defecto
            
            if sheet is None:
                messagebox.showerror("Error", "El archivo Excel no tiene una hoja activa.")
                return
            
            assert sheet is not None # Hint for type checker

            success_count = 0
            error_count = 0
            errors = []
            
            # Iterar filas empezando desde la 2 (saltar cabecera)
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Asumir columnas: A=Codigo, B=Nombre, C=DNI, D=Puesto, E=CentroCoste, F=Subdivision
                    if not row[0]:
                        continue
                        
                    codigo = str(row[0]).strip()
                    nombre = str(row[1]).strip() if len(row) > 1 and row[1] else f"Empleado {codigo}"
                    dni = str(row[2]).strip() if len(row) > 2 and row[2] else "00000000"
                    puesto = str(row[3]).strip() if len(row) > 3 and row[3] else "Sin Asignar"
                    centro_coste = str(row[4]).strip() if len(row) > 4 and row[4] else "1" # Default CC
                    subdivision = str(row[5]).strip() if len(row) > 5 and row[5] else None
                    
                    result = self.employee_service.create_employee(
                        codigo=codigo,
                        nombre=nombre,
                        dni=dni,
                        puesto=puesto,
                        codigo_centro_coste=centro_coste,
                        subdivision=subdivision
                    )
                    
                    if result.ok:
                        success_count += 1
                    else:
                        # Si falla, intentar actualizar? No, por ahora solo reportar
                        error_count += 1
                        if len(errors) < 5:
                            errors.append(f"Fila {idx} ({codigo}): {result.message}")
                            
                except Exception as e:
                    error_count += 1
                    if len(errors) < 5:
                        errors.append(f"Fila {idx}: Error inesperado - {str(e)}")
            
            msg = (f"Proceso completado.\n\n"
                   f"✅ Importados: {success_count}\n"
                   f"❌ Errores: {error_count}")
            
            if errors:
                msg += "\n\nDetalle de errores (primeros 5):\n" + "\n".join(errors)
            
            messagebox.showinfo("Resultado de Importación", msg)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al leer archivo:\n{str(e)}")

# -------------------------------------------------------------------------
    # MÉTODOS NUEVOS Y MODIFICADOS PARA IMPORTACIÓN DE 1 TRABAJADOR
    # -------------------------------------------------------------------------

    def _import_attendance_excel(self):
        """
        Importa asistencias analizando un reporte.
        Soporta dos formatos:
        1. Reporte Individual (Kardex): Código en cabecera.
        2. Reporte Detallado (Lista): Código en columna de tabla.
        """
        if not self.attendance_service:
            messagebox.showerror("Error", "No hay conexión a la base de datos")
            return

        if openpyxl is None:
            messagebox.showerror("Error", "La librería 'openpyxl' no está instalada.")
            return

        filename = filedialog.askopenfilename(
            title="Seleccionar Reporte de Asistencia",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.load_workbook(filename, data_only=True)
            sheet = wb.active

            if sheet is None:
                messagebox.showerror("Error", "El archivo Excel no tiene una hoja activa.")
                return
            
            assert sheet is not None

            # 1. DETECTAR ESTRUCTURA (Encabezados)
            headers_map, start_row = self._find_column_indexes(sheet)
            
            if headers_map['fecha'] is None:
                messagebox.showerror("Error de Formato", "No se encontró la columna 'Fecha' en el archivo.")
                return

            # 2. DETERMINAR ESTRATEGIA DE CÓDIGO
            # Si hay columna de código, usamos esa. Si no, buscamos en cabecera.
            col_idx_codigo = headers_map.get('codigo')
            codigo_header = None
            
            if col_idx_codigo is None:
                codigo_header = self._extract_employee_code(sheet)
                if not codigo_header:
                    from tkinter import simpledialog
                    codigo_header = simpledialog.askstring(
                        "Código no detectado", 
                        "No se detectó columna 'Código' ni cabecera.\nIngrese el código único para este archivo:"
                    )
                    if not codigo_header:
                        return

            # 3. PROCESAR FILAS
            success_count = 0
            error_count = 0
            skipped_count = 0
            errors = []
            processed_codes = set()

            for idx, row in enumerate(sheet.iter_rows(min_row=start_row + 1, values_only=True), start=start_row + 1):
                try:
                    # Determinar Código
                    if col_idx_codigo is not None:
                        val_code = row[col_idx_codigo]
                        current_code = str(val_code).strip() if val_code else None
                    else:
                        current_code = codigo_header

                    if not current_code:
                        continue
                        
                    processed_codes.add(current_code)

                    # Validar Fecha
                    fecha_val = row[headers_map['fecha']]
                    if not fecha_val:
                        continue

                    fecha_str, dia_semana = self._parse_excel_date(fecha_val)
                    if not fecha_str:
                        continue

                    # Validar Turno
                    codigo_turno = "GEN"
                    val_turno_raw = None
                    if headers_map['turno'] is not None:
                        val_turno = row[headers_map['turno']]
                        if val_turno:
                            val_turno_raw = str(val_turno)
                            codigo_turno = val_turno_raw.split(' ')[0][:10]

                    # Asegurar que el turno exista
                    self._ensure_shift_exists(codigo_turno, val_turno_raw)

                    # Validar Marcas
                    marca_entrada = None
                    marca_salida = None
                    
                    if headers_map['entrada'] is not None:
                        marca_entrada = self._parse_excel_time(row[headers_map['entrada']])
                    
                    if headers_map['salida'] is not None:
                        marca_salida = self._parse_excel_time(row[headers_map['salida']])

                    # CRITERIO: Solo importar si hay al menos una marca
                    if not marca_entrada and not marca_salida:
                        skipped_count += 1
                        continue

                    # Insertar
                    result = self.attendance_service.create_attendance(
                        fecha=fecha_str,
                        codigo_empleado=current_code,
                        codigo_turno=codigo_turno,
                        dia=dia_semana or "",
                        marca_entrada=marca_entrada,
                        marca_salida=marca_salida
                    )
                    
                    if result.ok:
                        success_count += 1
                    else:
                        error_count += 1
                        if len(errors) < 5:
                            errors.append(f"Fila {idx} ({current_code}): {result.message}")

                except Exception as e:
                    error_count += 1

            # 4. RESUMEN
            empleados_str = f"{len(processed_codes)} empleados detectados" if len(processed_codes) > 1 else f"Empleado: {list(processed_codes)[0] if processed_codes else '?'}"
            
            msg = (f"Importación Finalizada\n{empleados_str}\n\n"
                   f"✅ Registros creados: {success_count}\n"
                   f"⏭️ Omitidos (sin marcas): {skipped_count}\n"
                   f"❌ Errores: {error_count}")
            
            if errors:
                msg += "\n\nErrores (primeros 5):\n" + "\n".join(errors)
            
            messagebox.showinfo("Resultado", msg)

        except Exception as e:
            messagebox.showerror("Error Crítico", f"Error procesando el archivo:\n{str(e)}")

    def _extract_employee_code(self, sheet):
        """
        Busca en las primeras 15 filas alguna celda que diga 'Código', 'Legajo' o 'DNI'
        y toma el valor de la celda siguiente o adyacente.
        """
        keywords = ['código', 'codigo', 'legajo', 'trabajador', 'dni', 'id', 'cod.']
        
        # Escanear las primeras 15 filas y primeras 10 columnas
        for row in sheet.iter_rows(min_row=1, max_row=15, max_col=10, values_only=True):
            for i, cell_value in enumerate(row):
                if cell_value and isinstance(cell_value, str):
                    val_lower = cell_value.lower().strip()
                    
                    # Caso 1: La celda es exactamente la keyword (o con :)
                    # Ej: "Código" o "Código:"
                    clean_val = val_lower.replace(':', '').strip()
                    if clean_val in keywords:
                        # El valor está en la celda de la derecha (i+1)
                        if i + 1 < len(row) and row[i+1]:
                            return str(row[i+1]).strip()
                    
                    # Caso 2: El valor está en la misma celda
                    # Ej: "Código: E001" o "Legajo 1234"
                    for k in keywords:
                        if val_lower.startswith(k):
                            # Intentar separar por ':'
                            if ':' in cell_value:
                                parts = cell_value.split(':')
                                if len(parts) > 1:
                                    val = parts[1].strip()
                                    if val: return val
                            
                            # Si no tiene ':', ver si hay algo después de la keyword
                            # Ej: "Legajo 12345"
                            # Quitamos la keyword del inicio
                            remainder = val_lower[len(k):].strip()
                            if remainder:
                                # Retornamos la parte correspondiente del string original
                                # para preservar mayúsculas/minúsculas del código
                                # Ajuste: buscar la posición de la keyword en el original para cortar bien
                                start_idx = cell_value.lower().find(k)
                                if start_idx != -1:
                                    return cell_value[start_idx + len(k):].strip()
                            
        return None

    def _find_column_indexes(self, sheet):
        """
        Busca la fila de encabezados de la tabla y mapea las columnas.
        Retorna: (dict_map, row_index)
        """
        map_cols: Dict[str, Any] = {'fecha': None, 'turno': None, 'entrada': None, 'salida': None, 'codigo': None}
        found_header_row = 0
        
        # Palabras clave para identificar columnas
        keys_fecha = ['fecha', 'date']
        keys_turno = ['turno', 'horario']
        keys_in = ['entrada', 'ingreso', 'inicio', 'in']
        keys_out = ['salida', 'fin', 'out', 'retiro']
        keys_code = ['codigo', 'legajo', 'dni', 'trabajador']
        
        # Buscar hasta la fila 20
        for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            row_lower = [str(c).lower().strip() if c else '' for c in row]
            
            # Si encontramos "fecha" y ("entrada" o "ingreso"), es la fila de cabecera
            # Verificamos si alguna celda contiene alguna de las keywords
            has_fecha = any(any(k in cell for k in keys_fecha) for cell in row_lower)
            has_in = any(any(k in cell for k in keys_in) for cell in row_lower)
            has_out = any(any(k in cell for k in keys_out) for cell in row_lower)
            
            if has_fecha and (has_in or has_out):
                found_header_row = r_idx
                
                # Mapear índices (Priorizar la primera coincidencia encontrada de izquierda a derecha)
                for c_idx, val in enumerate(row_lower):
                    if map_cols['fecha'] is None and any(k in val for k in keys_fecha):
                        map_cols['fecha'] = c_idx
                    elif map_cols['turno'] is None and any(k in val for k in keys_turno):
                        map_cols['turno'] = c_idx
                    elif map_cols['entrada'] is None and any(k in val for k in keys_in):
                        map_cols['entrada'] = c_idx
                    elif map_cols['salida'] is None and any(k in val for k in keys_out):
                        map_cols['salida'] = c_idx
                    elif map_cols['codigo'] is None and any(k in val for k in keys_code):
                        map_cols['codigo'] = c_idx
                break
        
        # Fallback: Si no encuentra cabeceras claras, usar posiciones estándar de un Kardex común
        # Suponiendo: A=Fecha, B=Turno, C=Entrada, D=Salida
        if found_header_row == 0:
            # Asumir que los datos empiezan en fila 2 si no hay cabecera
            return {'fecha': 0, 'turno': 1, 'entrada': 2, 'salida': 3, 'codigo': None}, 1
            
        return map_cols, found_header_row

    def _parse_excel_date(self, val):
        """Auxiliar para convertir fecha de excel a string YYYY-MM-DD"""
        fecha_str = None
        dia_semana = ""
        
        try:
            if isinstance(val, datetime):
                fecha_str = val.strftime('%Y-%m-%d')
                dia_semana = self._get_day_name(val)
            elif isinstance(val, str):
                # Intentar parsear strings comunes
                for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d'):
                    try:
                        dt = datetime.strptime(val.strip(), fmt)
                        fecha_str = dt.strftime('%Y-%m-%d')
                        dia_semana = self._get_day_name(dt)
                        break
                    except ValueError:
                        continue
        except:
            pass
            
        return fecha_str, dia_semana

    def _parse_excel_time(self, val):
        """Auxiliar para convertir hora de excel a string HH:MM:SS"""
        if val is None or str(val).strip() in ['-', '', 'None', 'nan']:
            return None
        try:
            if isinstance(val, (datetime, time)):
                return val.strftime('%H:%M:%S')
            # Si es string, limpiarlo
            return str(val).strip()
        except:
            return None

    def _get_day_name(self, dt):
        days = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
        return days[dt.weekday()]

    def _ensure_shift_exists(self, shift_code, shift_raw_value=None):
        """
        Verifica si el turno existe en la BD. Si no, lo crea.
        Usa una caché local para evitar consultas repetitivas.
        """
        if not hasattr(self, '_shift_cache'):
            self._shift_cache = set()
            # Cargar turnos existentes
            if self.attendance_service and self.attendance_service.db:
                success, _, results = self.attendance_service.db.execute_procedure("sp_listar_turnos")
                if success:
                    self._shift_cache = {r['codigo_turno'] for r in results}
        
        if shift_code in self._shift_cache:
            return True
            
        # Si no está en caché, intentar insertarlo
        # Intentar parsear horarios del string raw (ej: "A10 (07:00-15:00)")
        start_time = '00:00:00'
        end_time = '00:00:00'
        
        if shift_raw_value and '(' in shift_raw_value and ')' in shift_raw_value:
            try:
                # Extraer contenido entre paréntesis
                times = shift_raw_value.split('(')[1].split(')')[0] # "07:00-15:00"
                if '-' in times:
                    parts = times.split('-')
                    if len(parts) == 2:
                        # Asegurar formato HH:MM:SS
                        t1 = parts[0].strip()
                        t2 = parts[1].strip()
                        start_time = t1 + ":00" if len(t1) == 5 else t1
                        end_time = t2 + ":00" if len(t2) == 5 else t2
            except:
                pass
                
        # Insertar en BD
        if self.attendance_service and self.attendance_service.db:
            query = "INSERT INTO turnos (codigo_turno, hora_entrada, hora_salida) VALUES (%s, %s, %s)"
            success, msg, _ = self.attendance_service.db.execute_insert(query, (shift_code, start_time, end_time))
            
            if success:
                self._shift_cache.add(shift_code)
                return True
            else:
                return False
        return False
