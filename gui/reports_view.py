"""
Vista del módulo de Reportes
Autor: Joaquin Armando Loaiza Cruz
Fecha: 2025-11-11
Descripción: Generación de reportes de horas extras usando procedimientos almacenados
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import Optional, Any, cast
from datetime import datetime, timedelta
import sys
import os

# Importar openpyxl para manejo de Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.report_service import ReportService
from database.employee_service import EmployeeService


class ReportsView:
    """Vista para generar reportes"""
    
    def __init__(self, parent_frame: tk.Frame,
                 report_service: Optional[ReportService],
                 employee_service: Optional[EmployeeService]):
        """
        Inicializa la vista de reportes.
        
        Args:
            parent_frame: Frame contenedor principal
            report_service: Servicio de reportes
            employee_service: Servicio de empleados
        """
        self.parent_frame = parent_frame
        self.report_service = report_service
        self.employee_service = employee_service
        
    def render(self):
        """Renderiza la vista completa de reportes"""
        # Limpiar frame padre
        for widget in self.parent_frame.winfo_children():
            widget.destroy()
        
        # Verificar conexión
        if not self.report_service:
            self._show_no_connection()
            return
        
        # Título
        self._create_header()
        
        # Contenedor principal
        main_container = tk.Frame(self.parent_frame, bg='#f5f5f5')
        main_container.pack(fill='both', expand=True, padx=30, pady=(0, 30))
        
        # Filtros y botones
        self._create_report_options(main_container)
        
        # Área de resultados
        self.results_container = tk.Frame(main_container, bg='#f5f5f5')
        self.results_container.pack(fill='both', expand=True, pady=(20, 0))
    
    def _show_no_connection(self):
        """Muestra mensaje de sin conexión"""
        error_frame = tk.Frame(self.parent_frame, bg='#ffebee', relief='solid', borderwidth=1)
        error_frame.pack(fill='both', expand=True, padx=30, pady=30)
        
        tk.Label(
            error_frame,
            text="⚠️ Sin Conexión a Base de Datos",
            font=('Segoe UI', 16, 'bold'),
            bg='#ffebee',
            fg='#c62828'
        ).pack(pady=(50, 10))
        
        tk.Label(
            error_frame,
            text="Por favor, conecte a la base de datos.",
            font=('Segoe UI', 11),
            bg='#ffebee',
            fg='#d32f2f'
        ).pack(pady=(0, 50))
    
    def _create_header(self):
        """Crea el encabezado"""
        title_frame = tk.Frame(self.parent_frame, bg='white', height=80)
        title_frame.pack(fill='x', padx=30, pady=(30, 20))
        title_frame.pack_propagate(False)
        
        title_content = tk.Frame(title_frame, bg='white')
        title_content.pack(side='left', expand=True)
        
        tk.Label(
            title_content,
            text="📈",
            font=('Segoe UI', 28),
            bg='white'
        ).pack(side='left', padx=(0, 15))
        
        tk.Label(
            title_content,
            text="Reportes de Horas Extras",
            font=('Segoe UI', 20, 'bold'),
            bg='white',
            fg='#2c3e50'
        ).pack(side='left')
    
    def _create_report_options(self, parent):
        """Crea las opciones de reporte"""
        options_frame = tk.Frame(parent, bg='white', relief='solid', borderwidth=1)
        options_frame.pack(fill='x')
        
        content = tk.Frame(options_frame, bg='white', padx=30, pady=20)
        content.pack(fill='x')
        
        tk.Label(
            content,
            text="📊 Configurar Reporte",
            font=('Segoe UI', 12, 'bold'),
            bg='white',
            fg='#2c3e50'
        ).grid(row=0, column=0, columnspan=4, sticky='w', pady=(0, 15))
        
        # Tipo de Reporte
        tk.Label(content, text="Tipo de Reporte:", font=('Segoe UI', 10), bg='white').grid(row=1, column=0, sticky='w', padx=(0, 10))
        self.tipo_reporte = tk.StringVar(value="empleado")
        
        tk.Radiobutton(
            content,
            text="Por Empleado",
            variable=self.tipo_reporte,
            value="empleado",
            font=('Segoe UI', 9),
            bg='white',
            command=self._toggle_employee_filter
        ).grid(row=1, column=1, sticky='w', padx=(0, 15))
        
        # Fecha Inicio
        tk.Label(content, text="Fecha Inicio:", font=('Segoe UI', 10), bg='white').grid(row=2, column=0, sticky='w', padx=(0, 10), pady=(10, 0))
        self.fecha_inicio_entry = tk.Entry(content, font=('Segoe UI', 10), width=15)
        self.fecha_inicio_entry.insert(0, (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        self.fecha_inicio_entry.grid(row=2, column=1, sticky='w', pady=(10, 0))
        
        # Fecha Fin
        tk.Label(content, text="Fecha Fin:", font=('Segoe UI', 10), bg='white').grid(row=2, column=2, sticky='w', padx=(15, 10), pady=(10, 0))
        self.fecha_fin_entry = tk.Entry(content, font=('Segoe UI', 10), width=15)
        self.fecha_fin_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        self.fecha_fin_entry.grid(row=2, column=3, sticky='w', pady=(10, 0))
        
        # Empleado (solo para reporte por empleado)
        tk.Label(content, text="Empleado:", font=('Segoe UI', 10), bg='white').grid(row=3, column=0, sticky='w', padx=(0, 10), pady=(10, 0))
        self.empleado_combo = ttk.Combobox(content, font=('Segoe UI', 10), state='readonly', width=30)
        
        if self.employee_service:
            employees = self.employee_service.get_all_employees() or []
            self.empleado_combo['values'] = ['Todos'] + [f"{e['codigo']} - {e['nombre']}" for e in employees]
            self.empleado_combo.current(0)
        
        self.empleado_combo.grid(row=3, column=1, columnspan=2, sticky='w', pady=(10, 0))
        
        # Botones
        btn_frame = tk.Frame(content, bg='white')
        btn_frame.grid(row=4, column=0, columnspan=4, pady=(20, 0), sticky='ew')
        
        tk.Button(
            btn_frame,
            text="📊 Generar Reporte",
            command=self._generate_report,
            font=('Segoe UI', 10, 'bold'),
            bg='#2196f3',
            fg='white',
            activebackground='#1976d2',
            relief='flat',
            cursor='hand2',
            padx=30,
            pady=12
        ).pack(side='left', padx=(0, 10))
        
        tk.Button(
            btn_frame,
            text="💾 Exportar Excel",
            command=self._export_excel,
            font=('Segoe UI', 10),
            bg='#4caf50',
            fg='white',
            activebackground='#45a049',
            relief='flat',
            cursor='hand2',
            padx=30,
            pady=12
        ).pack(side='left')

        # Botón Exportar Nómina
        tk.Button(
            btn_frame,
            text="💰 Exportar Nómina",
            command=self._export_payroll_excel,
            font=('Segoe UI', 10),
            bg='#ff9800',
            fg='white',
            activebackground='#f57c00',
            relief='flat',
            cursor='hand2',
            padx=30,
            pady=12
        ).pack(side='left', padx=(10, 0))
    
    def _toggle_employee_filter(self):
        """Habilita/deshabilita el filtro de empleado"""
        if self.tipo_reporte.get() == "empleado":
            self.empleado_combo.config(state='readonly')
        else:
            self.empleado_combo.config(state='disabled')
    
    def _generate_report(self):
        """Genera el reporte según los filtros"""
        # Limpiar resultados previos
        for widget in self.results_container.winfo_children():
            widget.destroy()
        
        if not self.report_service:
            messagebox.showerror("Error", "No hay conexión a la base de datos")
            return
        
        assert self.report_service is not None
        
        fecha_inicio = self.fecha_inicio_entry.get().strip()
        fecha_fin = self.fecha_fin_entry.get().strip()
        tipo = self.tipo_reporte.get()
        
        try:
            if tipo == "empleado":
                empleado = self.empleado_combo.get()
                codigo_emp = None if empleado == "Todos" else empleado.split(' - ')[0]
                data = self.report_service.get_overtime_by_employee(fecha_inicio, fecha_fin, codigo_emp)
            else:
                data = self.report_service.get_overtime_by_cost_center(fecha_inicio, fecha_fin)
            
            if not data:
                tk.Label(
                    self.results_container,
                    text="No se encontraron datos para el período seleccionado",
                    font=('Segoe UI', 12),
                    bg='#f5f5f5',
                    fg='#546e7a'
                ).pack(pady=50)
                return
            
            # Mostrar resultados
            self._display_report_results(data, tipo)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar reporte:\n{str(e)}")
    
    def _display_report_results(self, data, tipo):
        """Muestra los resultados del reporte"""
        # Canvas con scroll
        canvas = tk.Canvas(self.results_container, bg='#f5f5f5', highlightthickness=0)
        scrollbar = tk.Scrollbar(self.results_container, orient="vertical", command=canvas.yview)
        results_frame = tk.Frame(canvas, bg='white', relief='solid', borderwidth=1)

        results_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        window_id = canvas.create_window((0, 0), window=results_frame, anchor="nw")
        canvas.bind(
            "<Configure>",
            lambda e: canvas.itemconfigure(window_id, width=e.width)
        )
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Título
        tk.Label(
            results_frame,
            text=f"📋 Resultados - {'Por Empleado' if tipo == 'empleado' else 'Por Centro de Coste'}",
            font=('Segoe UI', 12, 'bold'),
            bg='white',
            fg='#2c3e50',
            padx=20,
            pady=15
        ).pack(fill='x', anchor='w')
        
        # Header
        header_frame = tk.Frame(results_frame, bg='#37474f', height=40)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        if tipo == "empleado":
            headers = [
                ("CÓDIGO", 0.12),
                ("NOMBRE", 0.28),
                ("HORAS 25%", 0.15),
                ("HORAS 35%", 0.15),
                ("HORAS 100%", 0.15),
                ("TOTAL EXTRAS", 0.15)
            ]
        else:
            headers = [
                ("CENTRO COSTE", 0.20),
                ("NOMBRE", 0.30),
                ("HORAS 25%", 0.125),
                ("HORAS 35%", 0.125),
                ("HORAS 100%", 0.125),
                ("TOTAL EXTRAS", 0.125)
            ]
        
        for header_text, width in headers:
            tk.Label(
                header_frame,
                text=header_text,
                font=('Segoe UI', 9, 'bold'),
                bg='#37474f',
                fg='white',
                anchor='w',
                padx=15
            ).place(relx=sum([h[1] for h in headers[:headers.index((header_text, width))]]),
                   rely=0, relwidth=width, relheight=1)
        
        # Datos
        total_h25 = 0
        total_h35 = 0
        total_h100 = 0
        
        for idx, row in enumerate(data):
            row_bg = '#f8f9fa' if idx % 2 == 0 else 'white'
            
            row_frame = tk.Frame(results_frame, bg=row_bg, height=40)
            row_frame.pack(fill='x')
            row_frame.pack_propagate(False)
            
            h25 = float(row.get('total_horas_25', 0) or 0)
            h35 = float(row.get('total_horas_35', 0) or 0)
            h100 = float(row.get('total_horas_100', 0) or 0)
            total = h25 + h35 + h100
            
            total_h25 += h25
            total_h35 += h35
            total_h100 += h100
            
            if tipo == "empleado":
                values = [
                    row.get('codigo_empleado', ''),
                    row.get('nombre_empleado', ''),
                    f"{h25:.1f}",
                    f"{h35:.1f}",
                    f"{h100:.1f}",
                    f"{total:.1f}"
                ]
            else:
                values = [
                    row.get('codigo_centro_coste', ''),
                    row.get('nombre_centro_coste', ''),
                    f"{h25:.1f}",
                    f"{h35:.1f}",
                    f"{h100:.1f}",
                    f"{total:.1f}"
                ]
            
            for i, (text, width) in enumerate(zip(values, [h[1] for h in headers])):
                tk.Label(
                    row_frame,
                    text=str(text),
                    font=('Segoe UI', 9),
                    bg=row_bg,
                    fg='#2c3e50',
                    anchor='w',
                    padx=15
                ).place(relx=sum([h[1] for h in headers[:i]]),
                       rely=0, relwidth=width, relheight=1)
        
        # Totales
        total_frame = tk.Frame(results_frame, bg='#e3f2fd', height=45)
        total_frame.pack(fill='x')
        total_frame.pack_propagate(False)
        
        tk.Label(
            total_frame,
            text="TOTALES:",
            font=('Segoe UI', 10, 'bold'),
            bg='#e3f2fd',
            fg='#1976d2',
            anchor='w',
            padx=15
        ).place(relx=0, rely=0, relwidth=0.40, relheight=1)
        
        offset = 0.40 if tipo == "empleado" else 0.50
        tk.Label(total_frame, text=f"{total_h25:.1f}", font=('Segoe UI', 10, 'bold'), bg='#e3f2fd', fg='#1976d2', anchor='w', padx=15).place(relx=offset, rely=0, relwidth=0.15, relheight=1)
        tk.Label(total_frame, text=f"{total_h35:.1f}", font=('Segoe UI', 10, 'bold'), bg='#e3f2fd', fg='#1976d2', anchor='w', padx=15).place(relx=offset+0.15, rely=0, relwidth=0.15, relheight=1)
        tk.Label(total_frame, text=f"{total_h100:.1f}", font=('Segoe UI', 10, 'bold'), bg='#e3f2fd', fg='#1976d2', anchor='w', padx=15).place(relx=offset+0.30, rely=0, relwidth=0.15, relheight=1)
        
        total_general = total_h25 + total_h35 + total_h100
        tk.Label(total_frame, text=f"{total_general:.1f}", font=('Segoe UI', 10, 'bold'), bg='#e3f2fd', fg='#1976d2', anchor='w', padx=15).place(relx=offset+0.45 if tipo == "empleado" else offset+0.37, rely=0, relwidth=0.15, relheight=1)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Guardar datos para exportar
        self.last_report_data = data
        self.last_report_type = tipo
    
    def _export_excel(self):
        """Exporta el reporte en formato XLSX con estilo"""
        if not hasattr(self, 'last_report_data') or not self.last_report_data:
            messagebox.showwarning("Advertencia", "Primero debe generar un reporte")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("Todos los archivos", "*.*")],
            initialfile=f"reporte_horas_extras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not filename:
            return

        try:
            self._build_xlsx_report(filename)
            messagebox.showinfo("Éxito", f"Reporte exportado exitosamente:\n{filename}")
        except ImportError:
            messagebox.showerror(
                "Dependencia faltante",
                "No se encontró la librería 'openpyxl'.\nEjecute 'pip install -r requirements.txt' para instalarla."
            )
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar:\n{str(e)}")

    def _build_xlsx_report(self, filename: str) -> None:
        """Construye el archivo XLSX aplicando los estilos solicitados"""
        from openpyxl import Workbook  # type: ignore
        from openpyxl.cell.cell import Cell  # type: ignore
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore

        settings = self._get_excel_layout()
        columns = settings['columns']
        group_field = settings['group_field']
        title = settings['title']

        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            raise ValueError('No se pudo crear la hoja de Excel')
        sheet.title = "Reporte"

        def _cell(row_idx: int, col_idx: int) -> Cell:
            return cast(Cell, sheet.cell(row=row_idx, column=col_idx))

        # Estilos base
        title_font = Font(name='Segoe UI', size=14, bold=True, color='1F2933')
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='37474F')
        thin_border = Border(
            left=Side(style='thin', color='D0D7DE'),
            right=Side(style='thin', color='D0D7DE'),
            top=Side(style='thin', color='D0D7DE'),
            bottom=Side(style='thin', color='D0D7DE')
        )

        # Título
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = _cell(1, 1)
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.row_dimensions[1].height = 32

        # Cabecera
        sheet.row_dimensions[2].height = 20
        for idx, col in enumerate(columns, start=1):
            header_cell = _cell(2, idx)
            header_cell.value = col['title']
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = thin_border
            sheet.column_dimensions[get_column_letter(idx)].width = col.get('width', 15)

        # Datos
        palette = ['E3F2FD', 'E8F5E9', 'FFF8E1', 'FFEBEE', 'F3E5F5', 'E0F2F1']
        current_color_index = -1
        current_group = None
        totals = {col['key']: 0.0 for col in columns if col.get('type') == 'number'}

        def _to_float(value: Optional[Any]) -> float:
            try:
                if value is None:
                    return 0.0
                if isinstance(value, (int, float)):
                    return float(value)
                return float(str(value).replace(',', '.'))
            except (ValueError, TypeError):
                return 0.0

        row_pointer = 3
        for row in self.last_report_data:
            group_value = row.get(group_field)
            if group_value != current_group:
                current_group = group_value
                current_color_index = (current_color_index + 1) % len(palette)

            row_fill = PatternFill('solid', fgColor=palette[current_color_index])
            hours_sum = sum(_to_float(row.get(key)) for key in settings['hour_keys'])

            for idx, col in enumerate(columns, start=1):
                key = col['key']
                data_cell = _cell(row_pointer, idx)

                if key == '__total__':
                    cell_value = hours_sum
                else:
                    cell_value = row.get(key, '')
                    if col.get('type') == 'number':
                        cell_value = _to_float(cell_value)

                data_cell.value = cell_value
                data_cell.fill = row_fill
                data_cell.border = thin_border
                if col.get('type') == 'number':
                    data_cell.number_format = '0.00'
                    totals[key] = totals.get(key, 0.0) + float(cell_value or 0)
                else:
                    data_cell.alignment = Alignment(horizontal='left', vertical='center')

            row_pointer += 1

        # Fila de totales
        total_row = row_pointer + 1
        label_cell = _cell(total_row, 1)
        label_cell.value = 'TOTALES:'
        label_cell.font = Font(name='Segoe UI', size=11, bold=True, color='0D47A1')
        label_cell.alignment = Alignment(horizontal='left', vertical='center')
        label_cell.fill = PatternFill('solid', fgColor='BBDEFB')
        label_cell.border = thin_border

        for idx, col in enumerate(columns[1:], start=2):
            summary_cell = _cell(total_row, idx)
            summary_cell.fill = PatternFill('solid', fgColor='BBDEFB')
            summary_cell.border = thin_border
            if col.get('type') == 'number':
                total_value = totals.get(col['key'], 0.0)
                summary_cell.value = round(total_value, 2)
                summary_cell.number_format = '0.00'
                summary_cell.font = Font(name='Segoe UI', size=11, bold=True, color='0D47A1')
            else:
                summary_cell.value = ''

        workbook.save(filename)

    def _get_excel_layout(self):
        """Define columnas y claves según el tipo de reporte"""
        if self.last_report_type == 'empleado':
            columns = [
                {'title': 'COD SAP', 'key': 'codigo_empleado', 'width': 14, 'type': 'text'},
                {'title': 'EMPLEADO', 'key': 'nombre_empleado', 'width': 32, 'type': 'text'},
                {'title': 'CC NOMINA', 'key': 'codigo_centro_coste', 'width': 15, 'type': 'text'},
                {'title': 'HORAS 25%', 'key': 'total_horas_25', 'width': 14, 'type': 'number'},
                {'title': 'HORAS 35%', 'key': 'total_horas_35', 'width': 14, 'type': 'number'},
                {'title': 'HORAS 100%', 'key': 'total_horas_100', 'width': 14, 'type': 'number'},
                {'title': 'TOTAL EXTRAS', 'key': '__total__', 'width': 16, 'type': 'number'}
            ]
            hour_keys = ['total_horas_25', 'total_horas_35', 'total_horas_100']
            title = 'RESUMEN SEGÚN MARCACIÓN - POR EMPLEADO'
            group_field = 'codigo_centro_coste'
        else:
            columns = [
                {'title': 'CENTRO COSTE', 'key': 'codigo_centro_coste', 'width': 16, 'type': 'text'},
                {'title': 'NOMBRE', 'key': 'nombre_centro_coste', 'width': 34, 'type': 'text'},
                {'title': 'HORAS 25%', 'key': 'total_horas_25', 'width': 14, 'type': 'number'},
                {'title': 'HORAS 35%', 'key': 'total_horas_35', 'width': 14, 'type': 'number'},
                {'title': 'HORAS 100%', 'key': 'total_horas_100', 'width': 14, 'type': 'number'},
                {'title': 'TOTAL EXTRAS', 'key': '__total__', 'width': 16, 'type': 'number'}
            ]
            hour_keys = ['total_horas_25', 'total_horas_35', 'total_horas_100']
            title = 'RESUMEN SEGÚN MARCACIÓN - POR CENTRO DE COSTE'
            group_field = 'codigo_centro_coste'

        return {
            'columns': columns,
            'hour_keys': hour_keys,
            'title': title,
            'group_field': group_field
        }

    def _export_payroll_excel(self):
        """Exporta reporte de nómina con formato específico (COD SAP, CC NOMINA, CANTIDAD)"""
        if not hasattr(self, 'last_report_data') or not self.last_report_data:
            messagebox.showwarning("Advertencia", "Primero debe generar un reporte")
            return

        if self.last_report_type != 'empleado':
            messagebox.showwarning("Advertencia", "Para exportar la nómina, debe generar un reporte 'Por Empleado'.")
            return

        if openpyxl is None:
            messagebox.showerror("Error", "La librería 'openpyxl' no está instalada.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=f"Nomina_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            title="Guardar Reporte de Nómina"
        )
        
        if not filename:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Carga Nómina"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4472C4")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Cabeceras
            headers = ["COD SAP", "CC NOMINA", "CANTIDAD"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Datos
            row_num = 2
            for row in self.last_report_data:
                cod_sap = row.get('codigo_empleado')
                if not cod_sap: continue

                # Convertir valores a float
                def _get_float(key):
                    val = row.get(key, 0)
                    if val is None: return 0.0
                    return float(val)

                h25 = _get_float('total_horas_25')
                h35 = _get_float('total_horas_35')
                h100 = _get_float('total_horas_100')

                # Fila H25 (Código 1250)
                if h25 > 0:
                    self._write_row(ws, row_num, cod_sap, "1250", h25, border)
                    row_num += 1
                
                # Fila H35 (Código 1252)
                if h35 > 0:
                    self._write_row(ws, row_num, cod_sap, "1252", h35, border)
                    row_num += 1
                    
                # Fila H100 (Código 1260)
                if h100 > 0:
                    self._write_row(ws, row_num, cod_sap, "1260", h100, border)
                    row_num += 1

            # Ajustar ancho columnas
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15

            wb.save(filename)
            messagebox.showinfo("Éxito", f"Archivo generado correctamente:\n{filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error inesperado:\n{str(e)}")

    def _write_row(self, ws, row_num, cod, cc, cant, border):
        """Escribe una fila en el Excel"""
        ws.cell(row=row_num, column=1, value=cod).border = border
        ws.cell(row=row_num, column=2, value=cc).border = border
        ws.cell(row=row_num, column=3, value=cant).border = border
